# -*- coding: utf-8 -*-


import os
import openpyxl

# Pasta onde os arquivos estão localizados
folder_path = r'C:\Users\caminhoquedoarquivo'

# Lista de arquivos na pasta
file_list = os.listdir(folder_path)

# Loop pelos arquivos na lista
for file_name in file_list:

    # Verifica se o nome do arquivo contém "Proforma_Retail"

    if "Proforma_Retail" in file_name:
        # Caminho completo do arquivo Excel
        file_path = os.path.join(folder_path, file_name)

        output_file_path = os.path.join(folder_path, file_name.replace(".xlsx", ".sql"))

        # Abre o arquivo Excel e executa o código correspondente
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            with open(output_file_path, 'a') as f:  # modificado de 'w' para 'a'
                for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    external_id = row[0]
                    if external_id is not None:
                        external_id = '{:.0f}'.format(external_id)
                        sql_command = "INSERT INTO e o comando = '{}');\n".format(
                            external_id)
                        f.write(sql_command)
            with open(output_file_path, 'r') as f:
                print(f.read())
            print("Conversão feita com sucesso!")
        finally:
            # Verifica se o arquivo está aberto e, se estiver, fecha sem salvar
            if 'workbook' in locals():
                workbook.close()

    if "Proforma_Corp" in file_name:
        # Caminho completo do arquivo Excel
        file_path = os.path.join(folder_path, file_name)

        output_file_path = os.path.join(folder_path, file_name.replace(".xlsx", ".sql"))

        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            with open(output_file_path, 'w') as f:
                for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    external_id = row[0]
                    if external_id is not None:
                        external_id = '{:.0f}'.format(external_id)
                        sql_command = "INSERT INTO e o comando = '{}');\n".format(
                            external_id)
                        f.write(sql_command)

            with open(output_file_path, 'r') as f:
                print(f.read())
                print("Conversão feita com sucesso!")

        finally:
            # Verifica se o arquivo está aberto e, se estiver, fecha sem salvar
            if 'workbook' in locals():
                workbook.close()

    if "Proforma_Sme" in file_name:
        # Caminho completo do arquivo Excel
        file_path = os.path.join(folder_path, file_name)
        output_file_path = os.path.join(folder_path, file_name.replace(".xlsx", ".sql"))

        # Abre o arquivo Excel e executa o código correspondente
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            with open(output_file_path, 'w') as f:
                for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    external_id = row[0]
                    if external_id is not None:
                        external_id = '{:.0f}'.format(external_id)
                        sql_command = "INSERT INTO e o comando = '{}');\n".format(
                            external_id)
                        f.write(sql_command)

            with open(output_file_path, 'r') as f:
                print(f.read())
                print("Conversão feita com sucesso!")

        finally:
            # Verifica se o arquivo está aberto e, se estiver, fecha sem salvar
            if 'workbook' in locals():
                workbook.close()

    if "Production_Sme" in file_name:
        # Caminho completo do arquivo Excel
        file_path = os.path.join(folder_path, file_name)

        # Abre o arquivo Excel e executa o código correspondente
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        output_file_path = os.path.join(folder_path, file_name.replace(".xlsx", ".sql"))
        with open(output_file_path, 'w') as f:
            for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                external_id = row[0]
                if external_id is not None:
                    external_id = '{:.0f}'.format(external_id)
                sql_command = "INSERT INTO e o comando ='{}');\n".format(
                    external_id)
                f.write(sql_command)
        with open(output_file_path, 'r') as f:
            print(f.read())
        print("Conversão feita com sucesso!")

    if "Production_Corp" in file_name:
        # Caminho completo do arquivo Excel
        file_path = os.path.join(folder_path, file_name)

        # Abre o arquivo Excel e executa o código correspondente
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        output_file_path = os.path.join(folder_path, file_name.replace(".xlsx", ".sql"))
        with open(output_file_path, 'w') as f:
            for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                external_id = row[0]
                if external_id is not None:
                    external_id = '{:.0f}'.format(external_id)
                sql_command = "INSERT INTO e o comando = '{}');\n".format(
                    external_id)
                f.write(sql_command)
        with open(output_file_path, 'r') as f:
            print(f.read())
        print("Conversão feita com sucesso!")


    # Verifica se o nome do arquivo contém "Production_Retail"
    elif "Production_Retail" in file_name:
        # Caminho completo do arquivo Excel
        file_path = os.path.join(folder_path, file_name)

        # Abre o arquivo Excel e executa o código correspondente
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        output_file_path = os.path.join(folder_path, file_name.replace(".xlsx", ".sql"))
        with open(output_file_path, 'w') as f:
            for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                external_id = row[0]
                if external_id is not None:
                    external_id = '{:.0f}'.format(external_id)
                sql_command = "INSERT INTO e o comando ='{}');\n".format(
                    external_id)
                f.write(sql_command)
        with open(output_file_path, 'r') as f:
            print(f.read())
        print("Conversão feita com sucesso!")


